#!/usr/bin/env python3
"""Eksportuje data/games.csv z powrotem do arkusza .xlsx.

Dzieki temu dane nigdy nie sa zamkniete w jednym formacie - w kazdej chwili
mozna wrocic do Excela. Arkusz ma zamrozony naglowek i wlaczone filtry.

Uruchomienie:
    python3 scripts/export_xlsx.py [plik-wyjsciowy.xlsx]
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Brak openpyxl. Zainstaluj: pip install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
GAMES = ROOT / "data" / "games.csv"

HEADERS = {
    "title": "Nazwa",
    "year": "Rok premiery",
    "status": "Status",
    "priority": "Priorytet",
    "owned": "Mam?",
    "platforms": "Platformy",
    "vr": "VR",
    "rating": "Ocena 1-5",
    "finished_year": "Rok ukończenia",
    "hype": "Hype 1-10",
    "review": "Recenzja",
    "blog": "Na blogu?",
    "tags": "Tagi",
    "notes": "Notatki",
}

WIDTHS = {"title": 42, "platforms": 22, "tags": 22, "notes": 30}
NUMERIC = {"year", "rating", "finished_year", "hype"}


def main() -> int:
    target = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "Kolekcja_gier_export.xlsx"

    with GAMES.open(encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        columns = list(reader.fieldnames or [])
        rows = list(reader)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Kolekcja gier"

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")

    for index, column in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=index, value=HEADERS.get(column, column))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
        sheet.column_dimensions[get_column_letter(index)].width = WIDTHS.get(column, 14)

    for row_index, row in enumerate(rows, start=2):
        for column_index, column in enumerate(columns, start=1):
            value = (row.get(column) or "").strip()
            if column in NUMERIC and value.isdigit():
                sheet.cell(row=row_index, column=column_index, value=int(value))
            else:
                sheet.cell(row=row_index, column=column_index, value=value)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"

    workbook.save(target)
    print(f"Zapisano {len(rows)} gier -> {target}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
